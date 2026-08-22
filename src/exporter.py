"""
exporter.py

Responsável por transformar a lista de livros em um DataFrame do pandas,
aplicar filtros, exportar os resultados (CSV/Excel) e exibir um resumo.
"""

import os

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PASTA_SAIDA = "output"
CAMINHO_CSV = os.path.join(PASTA_SAIDA, "livros.csv")
CAMINHO_XLSX = os.path.join(PASTA_SAIDA, "livros.xlsx")

# Largura (em "caracteres") de cada coluna na planilha Excel.
# Colunas não listadas aqui recebem uma largura padrão de 20.
LARGURAS_COLUNAS_EXCEL = {
    "titulo": 45,
    "preco": 12,
    "disponibilidade": 18,
    "rating": 10,
    "url": 50,
}


def criar_dataframe(livros):
    """Converte a lista de dicionários de livros em um pandas.DataFrame."""
    df = pd.DataFrame(livros)

    if df.empty:
        return df

    # Garante que a coluna de preço seja numérica para permitir filtros
    # e cálculos estatísticos (preço médio, mais barato, mais caro).
    df["preco"] = pd.to_numeric(df["preco"], errors="coerce")

    return df


def filtrar_por_preco(df, preco_maximo):
    """
    Retorna apenas os livros cujo preço seja menor ou igual a `preco_maximo`.

    Linhas com preço não numérico (NaN) são descartadas do resultado,
    pois não é possível compará-las de forma confiável.
    """
    if df.empty:
        return df

    df_filtrado = df[df["preco"].notna() & (df["preco"] <= preco_maximo)]
    return df_filtrado.reset_index(drop=True)


def garantir_pasta_saida():
    """Cria a pasta de saída (output/) caso ainda não exista."""
    os.makedirs(PASTA_SAIDA, exist_ok=True)


def exportar_csv(df, caminho=CAMINHO_CSV):
    """
    Exporta o DataFrame para um arquivo CSV pronto para ser aberto
    diretamente no Excel em configurações regionais de português/Brasil.

    Decisão técnica: o Excel em pt-BR usa "," como separador decimal e,
    por isso, trata "," como o separador de campos "de fato" ao importar
    CSV automaticamente — usar "," como separador de colunas faria o Excel
    confundir os dois. A solução padrão do mercado é usar ";" como
    separador de colunas e "," como separador decimal (sep=";", decimal=","),
    que é exatamente o que esta função faz. O encoding "utf-8-sig" adiciona
    um BOM no início do arquivo, o que faz o Excel reconhecer corretamente
    acentos e caracteres especiais (ex.: "ç", "é") em vez de exibi-los
    corrompidos.
    """
    garantir_pasta_saida()
    df.to_csv(
        caminho,
        index=False,
        encoding="utf-8-sig",
        sep=";",
        decimal=",",
    )
    print(f"Arquivo CSV salvo em: {caminho}")


def exportar_excel(df, caminho=CAMINHO_XLSX):
    """
    Exporta o DataFrame para um arquivo Excel (.xlsx) formatado como uma
    tabela legível: cabeçalho em destaque, colunas com largura adequada,
    preços em formato monetário, avaliações numéricas, URLs clicáveis,
    cabeçalho fixo ao rolar e autofiltro habilitado.

    Os dados do DataFrame não são alterados — apenas a apresentação
    visual da planilha gerada.
    """
    garantir_pasta_saida()

    if df.empty:
        # Ainda geramos um arquivo válido (só com cabeçalho) para não
        # quebrar o fluxo caso a coleta não retorne nenhum livro.
        df.to_excel(caminho, index=False, engine="openpyxl")
        print(f"Arquivo Excel salvo em: {caminho}")
        return

    with pd.ExcelWriter(caminho, engine="openpyxl") as escritor:
        df.to_excel(escritor, index=False, sheet_name="Livros")
        _formatar_planilha_excel(escritor.sheets["Livros"], df)

    print(f"Arquivo Excel salvo em: {caminho}")


def _formatar_planilha_excel(planilha, df):
    """
    Aplica a formatação visual na planilha já escrita pelo pandas:
    larguras de coluna, negrito no cabeçalho, formato numérico de preço
    e avaliação, hyperlinks nas URLs, cabeçalho congelado e uma tabela
    do Excel com estilo e autofiltro.
    """
    total_linhas = len(df) + 1  # +1 por causa da linha de cabeçalho
    total_colunas = len(df.columns)
    ultima_coluna = get_column_letter(total_colunas)

    # Largura de cada coluna (título mais largo, preço/rating mais estreitos).
    for posicao, nome_coluna in enumerate(df.columns, start=1):
        letra_coluna = get_column_letter(posicao)
        planilha.column_dimensions[letra_coluna].width = LARGURAS_COLUNAS_EXCEL.get(
            nome_coluna, 20
        )

    # Cabeçalho em negrito e centralizado.
    for celula_cabecalho in planilha[1]:
        celula_cabecalho.font = Font(bold=True)
        celula_cabecalho.alignment = Alignment(horizontal="center", vertical="center")

    indice_preco = _indice_da_coluna(df, "preco")
    indice_rating = _indice_da_coluna(df, "rating")
    indice_url = _indice_da_coluna(df, "url")

    for numero_linha in range(2, total_linhas + 1):
        if indice_preco:
            celula = planilha.cell(row=numero_linha, column=indice_preco)
            celula.number_format = '"£"#,##0.00'
            celula.alignment = Alignment(horizontal="right")

        if indice_rating:
            celula = planilha.cell(row=numero_linha, column=indice_rating)
            celula.number_format = "0"
            celula.alignment = Alignment(horizontal="center")

        if indice_url:
            celula = planilha.cell(row=numero_linha, column=indice_url)
            if celula.value:
                celula.hyperlink = celula.value
                celula.style = "Hyperlink"

    # Mantém o cabeçalho visível ao rolar a planilha para baixo.
    planilha.freeze_panes = "A2"

    # Transforma o intervalo em uma tabela do Excel: já inclui autofiltro
    # e permite ordenar/filtrar cada coluna diretamente no cabeçalho.
    referencia = f"A1:{ultima_coluna}{total_linhas}"
    tabela = Table(displayName="TabelaLivros", ref=referencia)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
    )
    planilha.add_table(tabela)


def _indice_da_coluna(df, nome_coluna):
    """Retorna o índice (1-based) de uma coluna do DataFrame, ou None se não existir."""
    if nome_coluna not in df.columns:
        return None
    return df.columns.get_loc(nome_coluna) + 1


def exibir_confirmacao_exportacao(df, caminho_csv=CAMINHO_CSV, caminho_xlsx=CAMINHO_XLSX):
    """
    Exibe uma mensagem clara confirmando onde os arquivos foram salvos,
    seguida de uma prévia curta dos dados (5 primeiras linhas) para
    conferência rápida, sem poluir o terminal com a tabela inteira.
    """
    print("\n===== EXPORTAÇÃO CONCLUÍDA =====\n")
    print("CSV:")
    print(f"  {caminho_csv}\n")
    print("Excel:")
    print(f"  {caminho_xlsx}\n")

    if not df.empty:
        print("Prévia dos dados (5 primeiras linhas):")
        print(df.head().to_string(index=False))
        print()

    print("=================================\n")


def _formatar_livro(linha):
    """Formata uma linha do DataFrame como 'Título (£preço)'."""
    if linha is None:
        return "N/A"
    return f"{linha['titulo']} (£{linha['preco']:.2f})"


def exibir_resumo(df_completo, df_filtrado=None):
    """
    Exibe no terminal um resumo estatístico da coleta:
    total de livros, total após filtro, preço médio, livro mais barato,
    livro mais caro e distribuição das avaliações.
    """
    print("\n===== RESUMO =====")
    print(f"Livros coletados: {len(df_completo)}")

    if df_filtrado is not None:
        print(f"Livros após filtro: {len(df_filtrado)}")

    if df_completo.empty:
        print("Nenhum dado disponível para calcular estatísticas.")
        return

    precos_validos = df_completo["preco"].dropna()
    if not precos_validos.empty:
        preco_medio = precos_validos.mean()
        linha_mais_barato = df_completo.loc[df_completo["preco"].idxmin()]
        linha_mais_caro = df_completo.loc[df_completo["preco"].idxmax()]

        print(f"Preço médio: £{preco_medio:.2f}")
        print(f"Livro mais barato: {_formatar_livro(linha_mais_barato)}")
        print(f"Livro mais caro: {_formatar_livro(linha_mais_caro)}")
    else:
        print("Não há preços válidos para calcular estatísticas.")

    if "rating" in df_completo.columns:
        distribuicao = df_completo["rating"].value_counts().sort_index()
        print("Distribuição das avaliações:")
        for estrelas, quantidade in distribuicao.items():
            print(f"  {estrelas} estrela(s): {quantidade} livro(s)")

    print("===================\n")